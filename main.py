import numpy as np
import torch
import torch.nn as nn
from torch.utils.data import DataLoader
import argparse
import pandas as pd

from model import SSRN, DCRN, LiEtAl, AAN

from loss import NormalizedCrossEntropy, ReverseCrossEntropy, ReverseCrossEntropy2
import data_generator
from draw_feature_map import draw_feature_map
from utils import evalution

import datetime

parser = argparse.ArgumentParser()
parser.add_argument('--batch_size', type=int, default=16, help='input batch size')
parser.add_argument('--max_iter', type=int, default=100, help='max training iterations')
parser.add_argument('--iters', type=int, default=10, help='Experiments iterations')
parser.add_argument('--lr', type=float, default=0.001, help='learning rate, default=0.001')
opt = parser.parse_args()

device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")


pathDict = {('./datasets/KSC.mat', './datasets/KSC_gt.mat'): (8, range(4,13,4))}
# pathDict = {('./datasets/KSC.mat', './datasets/KSC_gt.mat'): (24, range(4,13,4))}

# pathDict = {('./datasets/Salinas_corrected.mat', './datasets/Salinas_gt.mat'): (8, range(4,13,4))}
# pathDict = {('./datasets/Salinas_corrected.mat', './datasets/Salinas_gt.mat'): (24, range(1,4,1))}

# pathDict = {('./datasets/PaviaU1/PaviaU.mat', './datasets/PaviaU1/PaviaU_gt.mat'): (20, range(8,25,8))}
# pathDict = {('./datasets/PaviaU1/PaviaU.mat', './datasets/PaviaU1/PaviaU_gt.mat'): (52, range(8,25,8))}


# for count in range(1, 2):
for (dataset, gt) in pathDict.keys():
    for count in range(1, 3):
        day = datetime.datetime.now()
        day_str = day.strftime('%m_%d_%H_%M')

        (CL_num, NL_list) = pathDict[(dataset, gt)]
        Acc_total = []
        AA_total = []
        kappa_total = []
        alphaalpha_total = []
        model_best = {}

        for i in range(opt.iters):

            Acc_list = []
            AA_list = []
            kappa_list = []
            alphaalpha_list = []
            Loss = []
            data = data_generator.DataGenerator(dataset,
                                                gt,
                                                hwz = 3 ,
                                                CL_num= CL_num,
                                                NL_num= 0)
            data.start()
            for NL_num in NL_list:

                data.add_noise(NL_num - data.NL_num)
                data.getshape()




                model = AAN(input_channels=data.in_channel,
                                   patch_size=data.hwz * 2 + 1,
                                   n_classes=data.class_num).to(device)



                nce = NormalizedCrossEntropy(data.class_num).to(device)
                rce = ReverseCrossEntropy(data.class_num).to(device)

                ce = nn.CrossEntropyLoss().to(device)
                Optimizer = torch.optim.Adam(model.parameters(), lr = opt.lr)

                train, test = data.to_tensor()

                trainloader = DataLoader(train, batch_size= opt.batch_size, shuffle= True, drop_last= True)
                testloader = DataLoader(test, batch_size= 16, shuffle = True, drop_last= False)

                for epoch in range(opt.max_iter):
                    training_loss = 0
                    for (info, fake_label, true_label) in trainloader:
                        Optimizer.zero_grad()
                        info = info.unsqueeze(1)
                        info = info.to(device, torch.float32)
                        fake_label = fake_label.to(device, torch.long)
                        res, alpha = model(info)

                        score = torch.ones_like(fake_label)

                        loss = alpha * nce(res, fake_label) + (1 - alpha) * rce(res, fake_label)

                        training_loss += loss.item()
                        loss.backward()
                        Optimizer.step()

                    Loss.append(training_loss/ len(trainloader))
                    if epoch > 1 and abs(Loss[-1] - Loss[-2]) < 1e-4:
                        print('break at epoch {}'.format(epoch))
                        break
                Acc, aa, kappa, alphaalpha = evalution(testloader, model, data.class_num)
                Acc_list.append(Acc)
                AA_list.append(aa)
                kappa_list.append(kappa)
                alphaalpha_list.append(alphaalpha)
                if NL_num not in model_best:
                    model_best[NL_num] = (model, data, Acc)
                    print('Best model initialize to {:.2f} when NL_num is {}'.format(Acc * 100, NL_num))

                else:
                    if Acc > model_best[NL_num][2]:
                        Acc_before = model_best[NL_num][2]
                        model_best[NL_num] = (model, data, Acc)
                        print('Best model update from {:.2f} to {:.2f} when NL_num is {}'.format(Acc_before *100, Acc * 100, NL_num))

            Acc_total.append(Acc_list)
            AA_total.append(AA_list)
            kappa_total.append(kappa_list)
            alphaalpha_total.append(alphaalpha_list)
            print(Acc_total)
        print('OA of classification is {},\n AA of classification is {},\n kappa of classification is {}'.format(np.mean(Acc_total, 0) * 100,
                                                                                                          np.mean(AA_total, 0) * 100,
                                                                                                          np.mean(kappa_total, 0)))

        AA_mean = np.mean(np.mean(AA_total, 0), 1) * 100
        print('AA is {}'.format(AA_mean))

        AA_std  = []
        for i in range(len(AA_mean)):
            AA_temp = [AA_total[j][i] for j in range(len(AA_total))]
            AA_std.append(np.std(np.mean(AA_temp, 1)) * 100)
        print('OA std is {},\n AA std is {},\n kappa std is {}'.format(np.std(Acc_total, 0) * 100,
                                                                        AA_std,
                                                                        np.std(kappa_total, 0)))

        class_result = (np.mean(AA_total, 0) * 100).T
        oa_result = np.mean(Acc_total, 0) * 100
        oa_std_result = np.std(Acc_total, 0) * 100
        aa_result = AA_mean
        aa_std_result = AA_std
        kappa_result = np.mean(kappa_total, 0)
        kappa_std_result =  np.std(kappa_total, 0)
        alphaalpha_result = np.mean(alphaalpha_total, 0)

        oa_result = np.expand_dims(oa_result, 0)
        oa_std_result = np.expand_dims(oa_std_result, 0)
        aa_result = np.expand_dims(aa_result, 0)
        aa_std_result = np.expand_dims(aa_std_result, 0)
        kappa_result = np.expand_dims(kappa_result, 0)
        kappa_std_result = np.expand_dims(kappa_std_result, 0)
        alphaalpha_result = np.expand_dims(alphaalpha_result, 0)

        class_result = np.concatenate((class_result, oa_result), axis=0)
        class_result = np.concatenate((class_result, oa_std_result), axis=0)
        class_result = np.concatenate((class_result, aa_result), axis=0)
        class_result = np.concatenate((class_result, aa_std_result), axis=0)
        class_result = np.concatenate((class_result, kappa_result), axis=0)
        class_result = np.concatenate((class_result, kappa_std_result), axis=0)
        class_result = np.concatenate((class_result, alphaalpha_result), axis=0)

        result_path ='result/' + day_str + '.xlsx'
        data = pd.DataFrame(class_result)
        writer = pd.ExcelWriter(result_path)
        data.to_excel(writer, 'page_1', float_format='%.2f')

        writer.save()
        writer.close()

        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import *

        import warnings

        warnings.filterwarnings('ignore')

        wb = load_workbook(result_path)
        ws = wb['page_1']
        for i in range(1, 100):
            for j in range(1, 10):
                ws.cell(row=i, column=j).number_format = '0.00'
        wb.save(result_path)

        numnum = 1
        for key in model_best.keys():
            print('key:', key)

            draw_feature_map('salinas', model_best[key][0], model_best[key][1], day_str, numnum = numnum)
            numnum = numnum + 1


